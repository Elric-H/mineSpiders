import os
import openpyxl
import xlrd
from datetime import datetime
from typing import Any, Union


class ExcelReader:
    """统一封装xls和xlsx的读取操作"""

    def __init__(self, filepath):
        self.filepath = filepath
        _, ext = os.path.splitext(filepath)
        ext = ext.lower()

        if ext == '.xlsx':
            self.wb = openpyxl.load_workbook(filepath, data_only=True)
            self.file_type = 'xlsx'
        elif ext == '.xls':
            self.wb = xlrd.open_workbook(filepath)
            self.file_type = 'xls'
        else:
            raise ValueError(f"不支持的文件格式: {ext}")

    @property
    def sheetnames(self):
        if self.file_type == 'xlsx':
            return self.wb.sheetnames
        return self.wb.sheet_names()

    def get_sheet(self, sheet_name):
        if self.file_type == 'xlsx':
            return self.wb[sheet_name]
        return self.wb.sheet_by_name(sheet_name)

    def _convert_cell_ref(self, cell_ref: str) -> tuple:
        """将A1格式的单元格引用转换为(行号, 列号)"""
        col_letter = ''.join([c for c in cell_ref if c.isalpha()])
        row_num = int(''.join([c for c in cell_ref if c.isdigit()]))

        # 转换列字母为数字 (A=1, B=2, ..., Z=26, AA=27等)
        col_num = 0
        for i, c in enumerate(reversed(col_letter.upper())):
            col_num += (ord(c) - ord('A') + 1) * (26 ** i)

        return (row_num, col_num)

    def get_cell_value(self, sheet, cell_ref: str) -> Any:
        """统一获取单元格值的方法"""
        row_num, col_num = self._convert_cell_ref(cell_ref)

        if self.file_type == 'xlsx':
            # openpyxl的行列索引从1开始
            cell = sheet.cell(row=row_num, column=col_num)
            value = cell.value

            # 处理xlsx的日期格式
            if isinstance(value, datetime):
                return value.strftime('%Y-%m-%d')
            return value
        else:
            # xlrd的行列索引从0开始
            try:
                value = sheet.cell_value(row_num - 1, col_num - 1)

                # 处理xls的日期格式
                if sheet.cell_type(row_num - 1, col_num - 1) == xlrd.XL_CELL_DATE:
                    value = xlrd.xldate.xldate_as_datetime(value, self.wb.datemode)
                    return value.strftime('%Y-%m-%d')
                return value
            except IndexError:
                print(f"单元格 {cell_ref} 超出范围")
                return None
            except Exception as e:
                print(f"读取单元格 {cell_ref} 出错: {str(e)}")
                return None

    def get_row_values(self, sheet, row_num: int, col_letters: dict) -> dict:
        """获取一行中多个列的值"""
        values = {}
        for field, col_letter in col_letters.items():
            cell_ref = f'{col_letter}{row_num}'
            values[field] = self.get_cell_value(sheet, cell_ref)
        return values

    def check_end_marker(self, sheet, row_num: int, end_marker: str) -> bool:
        """检查是否到达结束标记"""
        if self.file_type == 'xlsx':
            value = sheet.cell(row=row_num, column=1).value  # 第一列
        else:
            try:
                value = sheet.cell_value(row_num - 1, 0)  # xls行号从0开始
            except IndexError:
                return True

        return not value or str(value).strip() == end_marker
