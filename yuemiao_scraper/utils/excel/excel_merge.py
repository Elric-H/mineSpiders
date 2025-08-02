import openpyxl
from datetime import datetime, date
from dateutil.parser import parse

SOURCE_FILE = './source.xlsx'
TARGET_FILE = './output_summary.xlsx'
OUTPUT_FILE = './合并文件.xlsx'
CONDITION_COLUMNS = 4
TIME_COLUMN = 5
DATE_COLUMNS = [5, 7]  # 只处理 E 和 G 列为日期


def format_date_value(value):
    if not value:
        return ""
    if isinstance(value, (datetime, date)):
        return value.strftime('%Y-%m-%d')
    if isinstance(value, str):
        try:
            dt = parse(value)
            return dt.strftime('%Y-%m-%d')
        except Exception as e:
            print(f"[警告] 日期解析失败（format）: '{value}' → {e}")
    return str(value)


def parse_date_safe(val):
    try:
        return parse(val).date()
    except Exception as e:
        print(f"[警告] 日期解析失败（parse）: '{val}' → {e}")
        return None


def read_data_from_workbook(wb):
    """仅读取指定列为日期格式，其余强制文本"""
    sheet = wb.worksheets[0]
    data = []

    for row_idx, row in enumerate(sheet.iter_rows(values_only=True), start=1):
        if not any(row):
            continue
        formatted_row = []
        for col_idx, cell in enumerate(row, 1):
            if col_idx in DATE_COLUMNS:
                formatted_cell = format_date_value(cell)
            else:
                formatted_cell = str(cell) if cell is not None else ""
            formatted_row.append(formatted_cell)
        data.append(formatted_row)
    print(f"[信息] 读取工作簿共 {len(data)} 行（含表头）")
    return data


def write_data_to_workbook(data, output_file):
    wb = openpyxl.Workbook()
    sheet = wb.active

    for row_idx, row in enumerate(data, 1):
        for col_idx, value in enumerate(row, 1):
            cell = sheet.cell(row=row_idx, column=col_idx, value=value)
            cell.number_format = '@'  # 设置为文本格式

    wb.save(output_file)
    print(f"[信息] 写入完成，共输出 {len(data)} 行 → {output_file}")


def merge_by_insertion(source_data, target_data):
    if not target_data:
        return source_data

    header = target_data[0]
    result = target_data[:]
    existing_keys = {tuple(row[:5]) for row in result[1:]}

    print(f"[信息] 目标数据现有 {len(result) - 1} 条（不含表头）")

    inserted_count = 0
    skipped_count = 0

    for row_index, src_row in enumerate(source_data[1:], start=2):
        key_4 = tuple(src_row[:CONDITION_COLUMNS])
        key_5 = tuple(src_row[:5])
        src_date = parse_date_safe(src_row[TIME_COLUMN - 1])

        if key_5 in existing_keys:
            print(f"[跳过] 第 {row_index} 行前五列重复，未插入：{key_5}")
            skipped_count += 1
            continue

        inserted = False
        for i in range(1, len(result)):
            tgt_row = result[i]
            tgt_key_4 = tuple(tgt_row[:CONDITION_COLUMNS])
            tgt_date = parse_date_safe(tgt_row[TIME_COLUMN - 1])

            if tgt_key_4 == key_4:
                if src_date and tgt_date and src_date < tgt_date:
                    result.insert(i, src_row)
                    print(f"[插入-中间] 行 {row_index} 插入第 {i + 1} 行：{key_5} < {tgt_date}")
                    existing_keys.add(key_5)
                    inserted = True
                    inserted_count += 1
                    break

        if not inserted:
            last_group_index = -1
            for i in range(1, len(result)):
                if tuple(result[i][:CONDITION_COLUMNS]) == key_4:
                    last_group_index = i
            if last_group_index != -1:
                result.insert(last_group_index + 1, src_row)
                print(f"[插入-分组尾] 行 {row_index} 插入到分组末尾（第 {last_group_index + 2} 行）: {key_4}")
            else:
                result.append(src_row)
                print(f"[插入-文件尾] 行 {row_index} 新分组，追加至文件末尾：{key_4}")
            existing_keys.add(key_5)
            inserted_count += 1

    print(f"[汇总] 插入 {inserted_count} 行，跳过 {skipped_count} 行")
    return result


def main():
    print("[开始] 读取文件...")
    source_wb = openpyxl.load_workbook(SOURCE_FILE, data_only=True)
    target_wb = openpyxl.load_workbook(TARGET_FILE, data_only=True)

    source_data = read_data_from_workbook(source_wb)
    target_data = read_data_from_workbook(target_wb)

    print("[开始] 执行合并逻辑...")
    merged_data = merge_by_insertion(source_data, target_data)

    print("[开始] 写入结果文件...")
    write_data_to_workbook(merged_data, OUTPUT_FILE)

    print("[完成] 合并任务完成")


if __name__ == "__main__":
    main()
    print("程序执行完毕！")
    input("按 Enter 键退出...")

