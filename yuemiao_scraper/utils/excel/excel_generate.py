import os
import openpyxl
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from datetime import datetime

from excel_reader import ExcelReader

# 配置项
CONFIG = {
    'input_directory': './input_files',  # 要遍历的目录路径
    'output_filename': 'output_summary.xlsx',  # 输出文件名
    'sheet_name': '汇总数据',  # 输出文件的sheet名称

    # 固定字段的单元格配置（每个文件只读取一次）
    'fixed_fields': {
        '姓名': 'B3',
        '变动后部门/单位': 'F3',
        '变动后岗位': 'B4'
    },

    # 动态字段的列配置（每行数据都会读取）
    'dynamic_columns': {
        '执行工资级别起算时间': 'B',
        '文号': 'C',
        '下文时间': 'D',
        '调整前工资级别': 'E',
        '调整后工资级别': 'F',
        '本次调资原因': 'G',
        '备注': 'H'
    },

    # 数据行起始和结束标记
    'data_start_row': 6,  # 数据开始行
    'end_marker': '审批意见',  # 数据结束标记

    # 输出文件的列顺序
    'output_columns': [
        '姓名',
        '员工编号',
        '变动后部门/单位',
        '变动后岗位',
        '执行工资级别起算时间',
        '文号',
        '下文时间',
        '调整前执行工资级别',
        '调整前执行工资档位',
        '调整后执行工资级别',
        '调整后执行工资档位',
        '本次调资原因',
        '备注'
    ]
}


def parse_employee_id(filename):
    """
    从文件名解析员工编号（预留逻辑）
    目前直接返回文件名作为员工编号，可根据实际需求修改
    """
    return os.path.splitext(filename)[0]
    # return None


def parse_salary_info(cell_value):
    """
    解析工资级别信息，如GM/01
    返回 (级别, 档位) 或 (None, None)
    """
    if not cell_value or '/' not in str(cell_value):
        return cell_value, None
    parts = str(cell_value).split('/')
    return parts[0], parts[1] if len(parts) > 1 else None


def process_excel_file(filepath):
    """
    处理单个Excel文件，提取所有有效数据
    """
    reader = ExcelReader(filepath)  # 使用导入的类
    filename = os.path.basename(filepath)
    employee_id = parse_employee_id(filename)

    all_data = []

    for sheet_name in reader.sheetnames:
        sheet = reader.get_sheet(sheet_name)

        # 读取固定字段的值
        fixed_data = {
            '员工编号': employee_id
        }

        for field, cell_ref in CONFIG['fixed_fields'].items():
            try:
                # 确保单元格引用格式正确
                if not isinstance(cell_ref, str) or not cell_ref[0].isalpha():
                    print(f"警告: 无效的单元格引用格式 {cell_ref}")
                    cell_value = None
                else:
                    # 使用reader的方法获取单元格值
                    cell_value = reader.get_cell_value(sheet, cell_ref)

            except Exception as e:
                print(f"读取 {field} @ {cell_ref} 出错: {str(e)}")
                cell_value = None

            fixed_data[field] = cell_value

        # 处理动态数据行
        row_num = CONFIG['data_start_row']
        while True:
            # 检查是否到达结束标记（使用reader的方法）
            if reader.check_end_marker(sheet, row_num, CONFIG['end_marker']):
                break

            # 创建当前行数据的副本（包含固定字段）
            row_data = fixed_data.copy()

            # 读取动态字段的值（使用reader的方法）
            dynamic_values = reader.get_row_values(sheet, row_num, CONFIG['dynamic_columns'])
            row_data.update(dynamic_values)

            # 解析调整前工资级别
            before_salary = row_data['调整前工资级别']
            before_level, before_grade = parse_salary_info(before_salary)
            row_data['调整前执行工资级别'] = before_level
            row_data['调整前执行工资档位'] = before_grade

            # 解析调整后工资级别
            after_salary = row_data['调整后工资级别']
            after_level, after_grade = parse_salary_info(after_salary)
            row_data['调整后执行工资级别'] = after_level
            row_data['调整后执行工资档位'] = after_grade

            # 删除原始的工资级别字段（不在最终输出中）
            del row_data['调整前工资级别']
            del row_data['调整后工资级别']

            all_data.append(row_data)
            row_num += 1

    return all_data


def generate_output_file(data):
    """
    生成汇总的Excel文件
    """
    wb = Workbook()
    ws = wb.active
    ws.title = CONFIG['sheet_name']

    # 写入表头
    for col_num, column_title in enumerate(CONFIG['output_columns'], 1):
        col_letter = get_column_letter(col_num)
        ws[f'{col_letter}1'] = column_title

    # 写入数据
    for row_num, row_data in enumerate(data, 2):
        for col_num, column_key in enumerate(CONFIG['output_columns'], 1):
            col_letter = get_column_letter(col_num)
            cell_value = row_data.get(column_key, '')

            # 处理日期格式
            if isinstance(cell_value, datetime):
                cell_value = cell_value.strftime('%Y-%m-%d')

            ws[f'{col_letter}{row_num}'] = cell_value

    # 自动调整列宽
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter  # 获取列字母
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2) * 1.2
        ws.column_dimensions[column].width = adjusted_width

    # 保存文件
    wb.save(CONFIG['output_filename'])
    print(f"汇总文件已生成: {CONFIG['output_filename']}")


def main():
    """
    主函数：遍历目录，处理所有Excel文件
    """
    start_time = datetime.now()
    print(f"开始处理，时间: {start_time.strftime('%Y-%m-%d %H:%M:%S')}")

    all_data = []
    processed_files = 0

    # 遍历目录下的所有文件
    for root, dirs, files in os.walk(CONFIG['input_directory']):
        for file in files:
            if file.endswith(('.xlsx', '.xls')):
                filepath = os.path.join(root, file)
                try:
                    print(f"开始处理文件: {file}")
                    file_data = process_excel_file(filepath)
                    all_data.extend(file_data)
                    processed_files += 1
                    print(f"已处理文件: {file}")
                except Exception as e:
                    print(f"处理文件 {file} 时出错: {str(e)}")

    if not all_data:
        print("没有找到可处理的数据")
        return

    # 生成汇总文件
    generate_output_file(all_data)

    end_time = datetime.now()
    duration = end_time - start_time
    print(f"处理完成，时间: {end_time.strftime('%Y-%m-%d %H:%M:%S')}")
    print(f"共处理 {processed_files} 个文件，生成 {len(all_data)} 条记录")
    print(f"总耗时: {duration.total_seconds():.2f} 秒")


if __name__ == '__main__':
    main()
    print("程序执行完毕！")
    input("按 Enter 键退出...")
